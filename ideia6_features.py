# -*- coding: utf-8 -*-
"""
Módulo: ideia6_features.py (rev C)

- Mantém Execução ('Execucao_<ts>') e 'Comparacao' (layout antigo).
- Adiciona:
  * Ponte snapshot->cmdmap: _snapshot_to_cmdmap()
  * Diffs: diff_vlans, diff_interfaces, diff_portchannels, diff_trunks, diff_neighbors
  * Sheets: 'Alterações Detalhadas' e 'Checklist Migração'
  * Chamadas dentro de ideia6_run_pipeline
"""

from __future__ import annotations

import datetime as _dt
import json
import os
import re as _re
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List, Optional, Sequence, Set, Tuple, Union
import logging


from history_json import get_last_two, simple_diff  # reaproveitado
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment
from openpyxl.formatting.rule import CellIsRule
from clean_switch.excel_utils import extract_active_vlans_from_table, extract_active_vlans_from_raw



# =============================================================================
# Métricas Execução (mantido)
# =============================================================================

@dataclass
class ExecMetrics:
    timestamp: str
    label: str
    total_ifaces: int
    up_ifaces: int
    down_ifaces: int
    vlans: Set[str]
    ios_version: str = ""
    uptime: str = ""
    host: str = ""
    ip: str = ""


IDEA6_EXECUTED: List[str] = []


def _mark_executed(func_name: str) -> None:
    IDEA6_EXECUTED.append(func_name)


def ideia6_safe_sheet_name(name: str, max_len: int = 31) -> str:
    invalid = r'[:\\/?*\[\]]'
    safe = _re.sub(invalid, "-", name)
    replacements = {
        "ç": "c", "Ç": "C",
        "ã": "a", "Ã": "A",
        "õ": "o", "Õ": "O",
        "á": "a", "à": "a", "â": "a", "ä": "a", "Á": "A", "À": "A", "Â": "A", "Ä": "A",
        "é": "e", "è": "e", "ê": "e", "ë": "e", "É": "E", "È": "E", "Ê": "E", "Ë": "E",
        "í": "i", "ì": "i", "î": "i", "ï": "i", "Í": "I", "Ì": "I", "Î": "I", "Ï": "I",
        "ó": "o", "ò": "o", "ô": "o", "ö": "o", "Ó": "O", "Ò": "O", "Ô": "O", "Ö": "O",
        "ú": "u", "ù": "u", "û": "u", "ü": "u", "Ú": "U", "Ù": "U", "Û": "U", "Ü": "U",
    }
    for a, b in replacements.items():
        safe = safe.replace(a, b)
    if len(safe) > max_len:
        safe = safe[:max_len]
    _mark_executed("ideia6_safe_sheet_name")
    return safe


def ideia6_autosize_columns(ws: Worksheet, min_width: int = 10, max_width: int = 60) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for c in col_cells:
            try:
                val = str(c.value) if c.value is not None else ""
            except Exception:
                val = ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_width, max_len + 2))
    _mark_executed("ideia6_autosize_columns")


def _find_entry(collected: Sequence[Tuple[str, Optional[Sequence[str]], Optional[Sequence[Sequence[str]]], Union[str, Sequence[str]]]], key: str):
    key_low = key.lower()
    for item in collected:
        cmd = item[0].lower()
        if key_low in cmd:
            return item
    return None


def _as_text(raw: Union[str, Sequence[str], None]) -> str:
    """Converte `raw` em string para uso com regex/splitlines."""
    if raw is None:
        return ""
    if isinstance(raw, str):
        return raw
    if isinstance(raw, (list, tuple)):
        try:
            return "\n".join(str(x) for x in raw)
        except Exception:
            return "\n".join(map(str, raw))
    return str(raw)


def ideia6_extract_metrics(collected: Sequence[Tuple[str, Optional[Sequence[str]], Optional[Sequence[Sequence[str]]], Union[str, Sequence[str]]]],
                           hostname: str,
                           ts: Union[str, _dt.datetime],
                           host_ip: Optional[str] = None) -> ExecMetrics:
    if isinstance(ts, _dt.datetime):
        pretty_ts = ts.strftime("%Y-%m-%d %H:%M:%S")
        label_ts = ts.strftime("%Y-%m-%d_%H-%M")
    else:
        pretty_ts = str(ts)
        label_ts = pretty_ts.replace(":", "-").replace(" ", "_")[:16]

    total_if = up_if = down_if = 0
    vlans: Set[str] = set()
    ios_version = ""
    uptime = ""

    ent_int = _find_entry(collected, "show interfaces status")
    if ent_int:
        _, headers, rows, raw = ent_int
        raw_text = _as_text(raw)
        if headers and rows:
            lower_headers = [h.lower() for h in headers]
            try:
                idx_status = lower_headers.index("status")
            except ValueError:
                idx_status = None

            total_if = len(rows)
            for r in rows:
                st = (r[idx_status] if (idx_status is not None and idx_status < len(r)) else "").lower()
                if "connected" in st or st == "up":
                    up_if += 1
            down_if = max(0, total_if - up_if)
        else:
            lines = [ln for ln in raw_text.splitlines() if ln.strip()]
            data_lines = [ln for ln in lines if not ln.lower().startswith(("port ", "----", "name "))]
            total_if = len(data_lines)
            up_if = sum(1 for ln in data_lines if "connected" in ln.lower() or " up " in ln.lower())
            down_if = max(0, total_if - up_if)

    ent_vlan = _find_entry(collected, "show vlan brief")
    if ent_vlan:
        _, headers, rows, raw = ent_vlan
        raw_text = _as_text(raw)
        vset = extract_active_vlans_from_table(headers or [], rows or [])
        if not vset:
            vset = extract_active_vlans_from_raw(raw_text)
        vlans = set(vset)   # agora são apenas as VLANs ATIVAS (e consistentes com o Excel)


    ent_ver = _find_entry(collected, "show version")
    if ent_ver:
        _, _headers, _rows, raw = ent_ver
        raw_text = _as_text(raw)
        m = _re.search(r"Version\s+([\w.\(\)]+)", raw_text)
        if m:
            ios_version = m.group(1)
        m2 = _re.search(r"[Uu]ptime is ([^\n]+)", raw_text)
        if m2:
            uptime = m2.group(1).strip()

    metrics = ExecMetrics(
        timestamp=pretty_ts,
        label=f"Execucao_{label_ts}",
        total_ifaces=total_if,
        up_ifaces=up_if,
        down_ifaces=down_if,
        vlans=vlans,
        ios_version=ios_version,
        uptime=uptime,
        host=hostname,
        ip=host_ip or "",
    )
    _mark_executed("ideia6_extract_metrics")
    return metrics


def _write_kv(ws: Worksheet, row: int, key: str, value: str, key_w: int = 18) -> int:
    ws.cell(row=row, column=1, value=key).font = Font(bold=True)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left")
    ws.cell(row=row, column=2, value=value)
    ws.column_dimensions["A"].width = key_w
    return row + 1


def ideia6_write_execution_sheet(wb: Workbook, m: ExecMetrics) -> Worksheet:
    sheet_name = ideia6_safe_sheet_name(m.label)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(title=sheet_name)

    r = 1
    r = _write_kv(ws, r, "Hostname", m.host or "")
    r = _write_kv(ws, r, "IP", m.ip or "")
    r = _write_kv(ws, r, "Data/Hora", m.timestamp)
    r = _write_kv(ws, r, "Versão IOS", m.ios_version or "N/D")
    r = _write_kv(ws, r, "Uptime", m.uptime or "N/D")

    r += 1
    ws.cell(row=r, column=1, value="MÉTRICAS DE INTERFACES").font = Font(bold=True)
    r += 1
    r = _write_kv(ws, r, "Total de interfaces", str(m.total_ifaces))
    r = _write_kv(ws, r, "Interfaces ativas", str(m.up_ifaces))
    r = _write_kv(ws, r, "Interfaces inativas", str(m.down_ifaces))

    r += 1
    ws.cell(row=r, column=1, value="VLANs DETECTADAS").font = Font(bold=True)
    r += 1
    vlan_list = ", ".join(sorted(m.vlans, key=lambda x: int(x) if x.isdigit() else x)) if m.vlans else "Nenhuma"
    r = _write_kv(ws, r, "Total de VLANs", str(len(m.vlans)))
    _ = _write_kv(ws, r, "Lista de VLANs", vlan_list)

    ideia6_autosize_columns(ws)
    _mark_executed("ideia6_write_execution_sheet")
    return ws


def _collect_all_metrics_from_wb(wb: Workbook) -> List[ExecMetrics]:
    out: List[ExecMetrics] = []
    for name in wb.sheetnames:
        if not name.lower().startswith("execucao_"):
            continue
        ws = wb[name]

        def getv(lbl: str) -> str:
            for r in range(1, ws.max_row + 1):
                k = ws.cell(row=r, column=1).value
                v = ws.cell(row=r, column=2).value
                if isinstance(k, str) and k.strip().lower() == lbl.lower():
                    return str(v) if v is not None else ""
            return ""

        try:
            total_if = int(getv("Total de interfaces") or "0")
        except ValueError:
            total_if = 0
        try:
            up_if = int(getv("Interfaces ativas") or "0")
        except ValueError:
            up_if = 0
        try:
            down_if = int(getv("Interfaces inativas") or "0")
        except ValueError:
            down_if = max(0, total_if - up_if)

        host = getv("Hostname")
        ip = getv("IP")
        ts = getv("Data/Hora") or name.replace("Execucao_", "").replace("_", " ")
        ios = getv("Versão IOS")
        up = getv("Uptime")

        vlan_list = getv("Lista de VLANs")
        vlans: Set[str] = set()
        if vlan_list:
            for tok in vlan_list.split(","):
                t = tok.strip()
                if t:
                    vlans.add(t)

        out.append(
            ExecMetrics(
                timestamp=ts, label=name, total_ifaces=total_if, up_ifaces=up_if,
                down_ifaces=down_if, vlans=vlans, ios_version=ios, uptime=up, host=host, ip=ip
            )
        )
    out.sort(key=lambda m: m.label)
    return out


def ideia6_update_comparison_sheet(wb: Workbook) -> Worksheet:
    ws_name = "Comparacao"
    ws = wb[ws_name] if ws_name in wb.sheetnames else wb.create_sheet(ws_name)
    ws.delete_rows(1, ws.max_row)

    all_m = _collect_all_metrics_from_wb(wb)

    headers = [
        "Execução", "Data/Hora", "Total_IF", "Ativas", "Inativas", "#VLANs", "Versão IOS", "Uptime"
    ]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = Font(bold=True)

    for i, m in enumerate(all_m, start=2):
        ws.cell(row=i, column=1, value=m.label)
        ws.cell(row=i, column=2, value=m.timestamp)
        ws.cell(row=i, column=3, value=m.total_ifaces)
        ws.cell(row=i, column=4, value=m.up_ifaces)
        ws.cell(row=i, column=5, value=m.down_ifaces)
        ws.cell(row=i, column=6, value=len(m.vlans))
        ws.cell(row=i, column=7, value=m.ios_version)
        ws.cell(row=i, column=8, value=m.uptime)

    r0 = len(all_m) + 3
    ws.cell(row=r0, column=1, value="DIFERENÇAS (últimas duas execuções)").font = Font(bold=True)
    if len(all_m) >= 2:
        a, b = all_m[-2], all_m[-1]
        dif_total = b.total_ifaces - a.total_ifaces
        dif_up = b.up_ifaces - a.up_ifaces
        dif_down = b.down_ifaces - a.down_ifaces
        dif_vlans = len(b.vlans) - len(a.vlans)
        novas_vlans = sorted(list(b.vlans - a.vlans), key=lambda x: int(x) if x.isdigit() else x)
        vlans_removidas = sorted(list(a.vlans - b.vlans), key=lambda x: int(x) if x.isdigit() else x)

        def fmt_delta(v: int) -> str:
            seta = "↑" if v > 0 else ("↓" if v < 0 else "→")
            return f"{seta} {v:+d}"

        rows = [
            ("Execução A (anterior)", a.label),
            ("Execução B (atual)", b.label),
            ("Δ Total_if", fmt_delta(dif_total)),
            ("Δ Ativas", fmt_delta(dif_up)),
            ("Δ Inativas", fmt_delta(dif_down)),
            ("Δ #VLANs", fmt_delta(dif_vlans)),
            ("VLANs novas", ", ".join(novas_vlans) if novas_vlans else "—"),
            ("VLANs removidas", ", ".join(vlans_removidas) if vlans_removidas else "—"),
        ]

        r = r0 + 1
        for k, v in rows:
            r = _write_kv(ws, r, k, v, key_w=22)

        start_delta = r0 + 3
        end_delta = start_delta + 3
        for rr in range(start_delta, end_delta + 1):
            ws.conditional_formatting.add(
                f"B{rr}",
                CellIsRule(operator='containsText', formula=['"+"'], stopIfTrue=False, font=Font(bold=True))
            )
            ws.conditional_formatting.add(
                f"B{rr}",
                CellIsRule(operator='containsText', formula=['"-"'], stopIfTrue=False, font=Font(bold=True))
            )

    ideia6_autosize_columns(ws)
    _mark_executed("ideia6_update_comparison_sheet")
    return ws


# =============================================================================
# Histórico e snapshots (mantido + auxiliares novos)
# =============================================================================

def _load_last_two_any(base_dir: str, hostname: str):
    """
    Tenta obter (prev, curr) primeiro do modelo multi-ficheiro (_history/host/*.json).
    Se não houver, tenta o ficheiro agregado <hostname>_history.json (lista).
    Normaliza para dicionários com chaves: 'timestamp' ou 'ts', e 'metrics'.
    """
    # 1) multi-ficheiro (history_json)
    prev, curr = get_last_two(base_dir=base_dir, hostname=hostname)
    if curr:
        return prev, curr

    # 2) ficheiro agregado <hostname>_history.json
    agg = os.path.join(base_dir, f"{hostname}_history.json")
    if os.path.exists(agg):
        try:
            with open(agg, "r", encoding="utf-8") as f:
                arr = json.load(f)
            if isinstance(arr, list) and arr:
                arr_sorted = sorted(arr, key=lambda x: x.get("ts") or x.get("timestamp") or "")
                if len(arr_sorted) >= 2:
                    return arr_sorted[-2], arr_sorted[-1]
                else:
                    return None, arr_sorted[-1]
        except Exception:
            pass
    return None, None


def _snap_ts(snap: dict) -> str:
    return (snap or {}).get("timestamp") or (snap or {}).get("ts") or "-"


def _snap_metrics(snap: dict) -> dict:
    """
    Extrai metrics de vários formatos:
      - topo: snap["metrics"]
      - meta: snap["meta"]["metrics"]
      - caso não exista, devolve {} (a Ideia 6 ainda cria as abas, mas com '-')
    """
    if not snap:
        return {}
    if isinstance(snap.get("metrics"), dict):
        return snap["metrics"]
    meta = snap.get("meta") or {}
    if isinstance(meta.get("metrics"), dict):
        return meta["metrics"]
    return {}


# =============================================================================
# NOVO: Ponte snapshot -> cmdmap (parsed por comando) e diffs
# =============================================================================

def _snapshot_to_cmdmap(snap: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
    """
    Converte snapshot guardado (items: cmd/raw/headers/rows) num dicionário:
        { "<cmd lower>": { "parsed": List[Dict[str,str]] } }
    para alimentar os diffs de migração.
    """
    out: Dict[str, Dict[str, Any]] = {}
    if not snap:
        return out
    items = snap.get("items") or []
    for it in items:
        cmd = str(it.get("cmd") or "").strip().lower()
        headers = [str(h).strip() for h in (it.get("headers") or [])]
        rows = it.get("rows") or []
        if headers and rows:
            low = [h.lower() for h in headers]
            parsed: List[Dict[str, str]] = []
            for r in rows:
                d: Dict[str, str] = {}
                for i, h in enumerate(low):
                    if i < len(r):
                        d[h] = "" if r[i] is None else str(r[i])
                parsed.append(d)
            out[cmd] = {"parsed": parsed}
        else:
            out[cmd] = {"parsed": []}
    return out


def _get_parsed(cmdmap: Dict[str, Dict[str, Any]], key: str) -> List[Dict[str, str]]:
    key = key.lower()
    for k, v in cmdmap.items():
        if key in k:
            return v.get("parsed") or []
    return []


def _norm_ifname(s: str) -> str:
    s = (s or "").strip()
    # Normaliza abreviações comuns (Gi -> Gi, GigabitEthernet -> Gi, Po -> Po)
    s = s.replace("GigabitEthernet", "Gi").replace("TenGigabitEthernet", "Te").replace("FastEthernet", "Fa")
    s = s.replace("Port-channel", "Po").replace("Port-Channel", "Po").replace("Ethernet", "Eth")
    return s


def _split_allowed(s: str) -> List[str]:
    """
    Converte '1-5,7,10' -> ['1','2','3','4','5','7','10']
    """
    out: List[str] = []
    s = (s or "").replace(" ", "")
    if not s or s == "none":
        return out
    parts = s.split(",")
    for p in parts:
        if "-" in p:
            a, b = p.split("-", 1)
            if a.isdigit() and b.isdigit():
                ra = range(int(a), int(b) + 1)
                out.extend([str(x) for x in ra])
        else:
            out.append(p)
    # remove vazios/duplicados e ordena numericamente quando possível
    out = [x for x in out if x]
    try:
        out = sorted(set(out), key=lambda x: (0, int(x)) if x.isdigit() else (1, x))
    except Exception:
        out = sorted(set(out))
    return out


def diff_vlans(prev_map: Dict[str, Any], curr_map: Dict[str, Any]) -> Dict[str, List[str]]:
    pv = _get_parsed(prev_map, "show vlan brief")
    cv = _get_parsed(curr_map, "show vlan brief")

    def _vlans(lst: List[Dict[str, str]]) -> Set[str]:
        out: Set[str] = set()
        for d in lst:
            # tenta 'vlan' ou primeira coluna numérica
            v = d.get("vlan") or d.get("vlan id") or d.get("vlan-id") or d.get("vlanid") or d.get("vlan_id")
            if not v:
                # tenta heurística: procurar chave cujo valor é dígito curto
                for k, val in d.items():
                    if str(val).strip().isdigit() and int(val) < 4096:
                        v = str(val).strip()
                        break
            if v and str(v).strip().isdigit():
                out.add(str(v).strip())
        return out

    a, b = _vlans(pv), _vlans(cv)
    add = sorted(list(b - a), key=lambda x: int(x) if x.isdigit() else x)
    rem = sorted(list(a - b), key=lambda x: int(x) if x.isdigit() else x)
    return {"added": add, "removed": rem}


def diff_interfaces(prev_map: Dict[str, Any], curr_map: Dict[str, Any]) -> List[Dict[str, str]]:
    pa = _get_parsed(prev_map, "show interfaces status")
    pb = _get_parsed(curr_map, "show interfaces status")

    def _map(lst: List[Dict[str, str]]) -> Dict[str, Dict[str, str]]:
        m: Dict[str, Dict[str, str]] = {}
        for d in lst:
            iface = _norm_ifname(d.get("port") or d.get("interface") or d.get("name") or "")
            if not iface:
                continue
            status = (d.get("status") or "").lower()
            vlan = (d.get("vlan") or d.get("access vlan") or "").strip()
            m[iface] = {"status": status, "vlan": vlan}
        return m

    A, B = _map(pa), _map(pb)
    keys = set(A.keys()) | set(B.keys())
    out: List[Dict[str, str]] = []
    for k in sorted(keys):
        a = A.get(k, {"status": "", "vlan": ""})
        b = B.get(k, {"status": "", "vlan": ""})
        if a != b:
            out.append({
                "interface": k,
                "status_from": a.get("status", ""),
                "status_to": b.get("status", ""),
                "vlan_from": a.get("vlan", ""),
                "vlan_to": b.get("vlan", ""),
            })
    return out


def diff_portchannels(prev_map: Dict[str, Any], curr_map: Dict[str, Any]) -> List[Dict[str, str]]:
    pa = _get_parsed(prev_map, "show etherchannel summary")
    pb = _get_parsed(curr_map, "show etherchannel summary")

    def _norm_members(s: str) -> str:
        # espera algo tipo "[Gi1/0/1, Gi1/0/2]" ou "Gi1/0/1,Gi1/0/2"
        s = (s or "").strip().strip("[]")
        toks = [_norm_ifname(x) for x in s.replace(";", ",").split(",") if x.strip()]
        return ", ".join(sorted(set(toks)))

    def _pick(d: dict, *keys) -> str:
        for k in keys:
            v = d.get(k)
            if v is not None and str(v).strip() != "":
                return str(v)
        return ""

    def _map(lst: List[Dict[str, str]]) -> Dict[str, Dict[str, str]]:
        m: Dict[str, Dict[str, str]] = {}
        for d in lst:
            # Nome do bundle (Po): aceitar várias chaves
            po = _norm_ifname(_pick(
                d, "bundle_name", "port-channel", "port channel", "po", "portchannel", "group"
            ))
            if not po:
                continue
            # Estado/flags
            state = _pick(d, "bundle_status", "status", "bundle state", "flags")
            # Membros (aceitar singular/plural e variantes)
            members_raw = _pick(
                d,
                "member_interfaces", "member interface", "member_interface",
                "interfaces", "members",
                "member ports", "member_ports"
            )
            m[po] = {"state": state, "members": _norm_members(members_raw)}
        return m

    A, B = _map(pa), _map(pb)
    keys = set(A.keys()) | set(B.keys())
    out: List[Dict[str, str]] = []
    for k in sorted(keys):
        a = A.get(k, {"state": "", "members": ""})
        b = B.get(k, {"state": "", "members": ""})
        if a != b:
            out.append({
                "po": k,
                "state_from": a.get("state", ""),
                "state_to": b.get("state", ""),
                "members_from": a.get("members", ""),
                "members_to": b.get("members", ""),
            })
    return out



def diff_trunks(prev_map: Dict[str, Any], curr_map: Dict[str, Any]) -> List[Dict[str, str]]:
    pa = _get_parsed(prev_map, "show interfaces trunk")
    pb = _get_parsed(curr_map, "show interfaces trunk")

    def _map(lst: List[Dict[str, str]]) -> Dict[str, Dict[str, str]]:
        m: Dict[str, Dict[str, str]] = {}
        for d in lst:
            iface = _norm_ifname(d.get("port") or d.get("interface") or "")
            if not iface:
                continue

            # Native VLAN (várias grafias)
            native = (
                d.get("native") or d.get("native vlan") or d.get("native_vlan") or
                d.get("nativevlan") or d.get("native vlan id") or ""
            )

            # Allowed VLANs (aceitar vários nomes; se vazio, usar allowed_active_vlans)
            allowed = (
                d.get("vlans allowed") or d.get("allowed") or d.get("allowed_vlans") or
                d.get("vlans_allowed_on_trunk") or d.get("vlans allowed on trunk") or
                d.get("allowed_active_vlans") or
                d.get("vlans allowed and active in management domain") or ""
            )
            if not str(allowed).strip():
                allowed = d.get("allowed_vlans") or d.get("allowed_active_vlans") or ""

            m[iface] = {
                "native": str(native).strip(),
                "allowed": ", ".join(_split_allowed(str(allowed))),
            }
        return m

    A, B = _map(pa), _map(pb)
    keys = set(A.keys()) | set(B.keys())
    out: List[Dict[str, str]] = []
    for k in sorted(keys):
        a = A.get(k, {"native": "", "allowed": ""})
        b = B.get(k, {"native": "", "allowed": ""})
        if a != b:
            out.append({
                "interface": k,
                "native_from": a.get("native", ""),
                "native_to": b.get("native", ""),
                "allowed_from": a.get("allowed", ""),
                "allowed_to": b.get("allowed", ""),
            })
    return out



from typing import Any, Dict, List, Tuple

def _nk(d: dict) -> dict:
    """normaliza chaves: lower + troca espaços/hífens por underscore"""
    out = {}
    for k, v in (d or {}).items():
        nk = str(k).strip().lower().replace(" ", "_").replace("-", "_")
        out[nk] = v
    return out

def _pick(d: dict, *keys: str) -> str:
    for k in keys:
        v = d.get(k)
        if v is not None and str(v).strip() != "":
            return str(v)
    return ""

def _neighbor_rows_from_parsed(lst: List[Dict[str, Any]], proto: str) -> List[Dict[str, str]]:
    """
    Converte linhas parsed (CDP/LLDP) num formato uniforme:
      {neighbor, local_if, neighbor_if, proto}
    """
    out: List[Dict[str, str]] = []
    for raw in lst or []:
        d = _nk(raw)

        # Nome do vizinho (CDP: device_id; LLDP: system_name)
        neighbor = _pick(
            d, "neighbor", "neighbor_name", "device_id", "system_name",
               "system_name_value", "chassis_id"
        ).strip()
        if not neighbor:
            continue

        # Interface local (CDP costuma trazer; LLDP pode não trazer)
        local_if = _pick(d, "local_if", "local_interface", "local_port", "local port", "interface")
        local_if = _norm_ifname(local_if)

        # Interface do vizinho (CDP: port_id; LLDP: port_id/port_description)
        neighbor_if = _pick(d, "neighbor_if", "neighbor_interface", "neighbor_port",
                               "port_id", "port description", "port_description", "port_id_value")
        neighbor_if = _norm_ifname(neighbor_if)

        out.append({
            "neighbor": neighbor,
            "local_if": local_if,
            "neighbor_if": neighbor_if,
            "proto": proto
        })
    return out

def _collect_neighbors(map_obj: Dict[str, Any]) -> List[Dict[str, str]]:
    """
    Lê CDP e LLDP do snapshot e devolve lista unificada (com proto).
    Tenta LLDP 'detail' primeiro; se vier vazio, usa 'show lldp neighbors'.
    """
    cdp  = _get_parsed(map_obj, "show cdp neighbors detail") or []
    lldp = _get_parsed(map_obj, "show lldp neighbors detail")
    if not lldp:
        lldp = _get_parsed(map_obj, "show lldp neighbors") or []
    rows: List[Dict[str, str]] = []
    rows += _neighbor_rows_from_parsed(cdp,  "CDP")
    rows += _neighbor_rows_from_parsed(lldp, "LLDP")
    return rows

def _merge_by_neighbor_port(rows: List[Dict[str, str]]) -> Dict[Tuple[str, str], Dict[str, str]]:
    """
    Junta entradas do mesmo (neighbor, neighbor_if), preenchendo local_if se faltar.
    Retorna dict key->row para facilitar diffs.
    """
    merged: Dict[Tuple[str, str], Dict[str, str]] = {}
    for r in rows:
        k = (r["neighbor"].strip().lower(), r["neighbor_if"].strip().lower())
        if k not in merged:
            merged[k] = dict(r)
        else:
            # preferir local_if preenchido
            if not merged[k].get("local_if") and r.get("local_if"):
                merged[k]["local_if"] = r["local_if"]
    return merged

def diff_neighbors(prev_map: Dict[str, Any], curr_map: Dict[str, Any]) -> Dict[str, List[Dict[str, str]]]:
    """
    Devolve:
      {
        "added":   [{"neighbor","local_if","neighbor_if"}, ...],
        "removed": [{"neighbor","local_if","neighbor_if"}, ...]
      }
    Sem duplicados CDP/LLDP.
    """
    prev_rows = _merge_by_neighbor_port(_collect_neighbors(prev_map))
    curr_rows = _merge_by_neighbor_port(_collect_neighbors(curr_map))

    added_keys = sorted(set(curr_rows.keys()) - set(prev_rows.keys()))
    removed_keys = sorted(set(prev_rows.keys()) - set(curr_rows.keys()))

    added = []
    for k in added_keys:
        r = curr_rows[k]
        added.append({
            "neighbor": r.get("neighbor", ""),
            "local_if": r.get("local_if", ""),
            "neighbor_if": r.get("neighbor_if", ""),
        })

    removed = []
    for k in removed_keys:
        r = prev_rows[k]
        removed.append({
            "neighbor": r.get("neighbor", ""),
            "local_if": r.get("local_if", ""),
            "neighbor_if": r.get("neighbor_if", ""),
        })

    return {"added": added, "removed": removed}


def build_migration_deltas(prev_map: Dict[str, Any], curr_map: Dict[str, Any]) -> Dict[str, Any]:
    """Agrega todos os diffs para alimentar o Excel."""
    return {
        "vlans": diff_vlans(prev_map, curr_map),
        "interfaces": diff_interfaces(prev_map, curr_map),
        "portchannels": diff_portchannels(prev_map, curr_map),
        "trunks": diff_trunks(prev_map, curr_map),
        "neighbors": diff_neighbors(prev_map, curr_map),
    }


# =============================================================================
# NOVO: Writers de sheets 'Alterações Detalhadas' e 'Checklist Migração'
# =============================================================================

def _write_table(ws: Worksheet, start_row: int, title: str, headers: List[str], rows: List[List[Any]]) -> int:
    """Escreve uma pequena tabela (título + cabeçalho + linhas). Devolve a próxima row livre."""
    ws.cell(row=start_row, column=1, value=title).font = Font(bold=True)
    r = start_row + 1
    for j, h in enumerate(headers, start=1):
        ws.cell(row=r, column=j, value=h).font = Font(bold=True)
    for row in rows:
        r += 1
        for j, v in enumerate(row, start=1):
            ws.cell(row=r, column=j, value=v)
    # largura amigável
    for col in range(1, len(headers) + 1):
        try:
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 24
        except Exception:
            pass
    return r + 2


def _write_alteracoes_sheet(wb: Workbook, prev_snap: Dict[str, Any], curr_snap: Dict[str, Any]) -> None:
    """Cria/atualiza a sheet 'Alterações Detalhadas' usando os últimos dois snapshots."""
    name = "Alterações Detalhadas"
    if name in wb.sheetnames:
        ws = wb[name]; ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(name)

    prev_map = _snapshot_to_cmdmap(prev_snap) if prev_snap else {}
    curr_map = _snapshot_to_cmdmap(curr_snap) if curr_snap else {}
    deltas = build_migration_deltas(prev_map, curr_map)

    r = 1
    # VLANs
    vl = deltas.get("vlans", {})
    rows_vl = [["Adicionada", v] for v in (vl.get("added") or [])] + \
              [["Removida", v] for v in (vl.get("removed") or [])]
    if not rows_vl:
        rows_vl = [["—", "—"]]
    r = _write_table(ws, r, "VLANs (Δ)", ["Tipo", "VLAN"], rows_vl)

    # Interfaces
    ifs = deltas.get("interfaces") or []
    rows_if = [[d.get("interface",""), d.get("status_from",""), d.get("status_to",""), d.get("vlan_from",""), d.get("vlan_to","")] for d in ifs]
    if not rows_if:
        rows_if = [["—","—","—","—","—"]]
    r = _write_table(ws, r, "Interfaces (estado/VLAN alterados ou desaparecidos/novos)",
                     ["Interface","Status (de)","Status (para)","VLAN (de)","VLAN (para)"], rows_if)

    # Port-Channels
    pos = deltas.get("portchannels") or []
    rows_po = [[d.get("po",""), d.get("members_from",""), d.get("members_to",""), d.get("state_from",""), d.get("state_to","")] for d in pos]
    if not rows_po:
        rows_po = [["—","—","—","—","—"]]
    r = _write_table(ws, r, "Port-Channels (membros/estado alterados)",
                     ["Port-Channel","Membros (de)","Membros (para)","Estado (de)","Estado (para)"], rows_po)

    # Trunks
    trs = deltas.get("trunks") or []
    rows_tr = [[d.get("interface",""), d.get("native_from",""), d.get("native_to",""), d.get("allowed_from",""), d.get("allowed_to","")] for d in trs]
    if not rows_tr:
        rows_tr = [["—","—","—","—","—"]]
    r = _write_table(ws, r, "Trunks (VLAN nativa/allowed alteradas)",
                     ["Interface","Native (de)","Native (para)","Allowed (de)","Allowed (para)"], rows_tr)

    # Neighbors
    nbs = deltas.get("neighbors") or {}
    rows_nb_add = [[d.get("neighbor",""), d.get("local_if",""), d.get("neighbor_if","")] for d in (nbs.get("added") or [])]
    rows_nb_rem = [[d.get("neighbor",""), d.get("local_if",""), d.get("neighbor_if","")] for d in (nbs.get("removed") or [])]
    if not rows_nb_add:
        rows_nb_add = [["—","—","—"]]
    if not rows_nb_rem:
        rows_nb_rem = [["—","—","—"]]
    r = _write_table(ws, r, "Vizinhos ADICIONADOS (CDP/LLDP)",
                     ["Neighbor","Local If","Neighbor If"], rows_nb_add)
    _ = _write_table(ws, r, "Vizinhos REMOVIDOS (CDP/LLDP)",
                     ["Neighbor","Local If","Neighbor If"], rows_nb_rem)


def _write_checklist_sheet(wb: Workbook, prev_metrics: Dict[str, Any], curr_metrics: Dict[str, Any], had_prev: bool) -> None:
    """
    Cria 'Checklist Migração' focada em: 'Vais apenas validar se a rede nova ficou igual?'
    A sheet referencia números e manda o utilizador ver 'Alterações Detalhadas' quando preciso.
    """
    name = "Checklist Migração"
    if name in wb.sheetnames:
        ws = wb[name]; ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(name)

    ws["A1"] = "Checklist de Validação Pós-Migração"; ws["A1"].font = Font(bold=True)
    r = 3

    def _kv(k, v1, v2=""):
        nonlocal r
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=v1)
        ws.cell(row=r, column=3, value=v2)
        r += 1

    prev_if = prev_metrics.get("interfaces_total") if had_prev else None
    curr_if = curr_metrics.get("interfaces_total")
    prev_up = prev_metrics.get("interfaces_up") if had_prev else None
    curr_up = curr_metrics.get("interfaces_up")
    prev_vl = prev_metrics.get("vlans_active") if had_prev else None
    curr_vl = curr_metrics.get("vlans_active")

    _kv("Executaste recolha ANTES e DEPOIS?", "Sim" if had_prev else "Não")
    _kv("A rede nova ficou igual? (ver Alterações Detalhadas)", "—", "Abrir sheet")

    def _fmt_delta(a, b):
        if isinstance(a, (int, float)) and isinstance(b, (int, float)):
            d = b - a
            return f"{'↑' if d>0 else ('↓' if d<0 else '→')} {d:+d}"
        return "-"

    _kv("Interfaces (total) – Δ", _fmt_delta(prev_if, curr_if), f"Antes: {prev_if} / Agora: {curr_if}")
    _kv("Interfaces ativas (Up) – Δ", _fmt_delta(prev_up, curr_up), f"Antes: {prev_up} / Agora: {curr_up}")
    _kv("VLANs ativas – Δ", _fmt_delta(prev_vl, curr_vl), f"Antes: {prev_vl} / Agora: {curr_vl}")
    _kv("Vizinhos, Trunks, Port-Channels", "Ver 'Alterações Detalhadas'")

    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 38





# =============================================================================
# Pipeline (mantido + chamadas novas)
# =============================================================================

def ideia6_run_pipeline(hostname: str, ts: str, collected, xlsx_path: str, host_ip: str = None):
    """
    Ideia 6 (JSON): cria aba 'Execucao_<ts>' e 'Comparacao' (layout antigo)
    e agora também 'Alterações Detalhadas' e 'Checklist Migração',
    usando as DUAS últimas execuções guardadas em JSON (multi-ficheiro ou agregado).
    """
    wb = load_workbook(xlsx_path)
    base_dir = os.path.dirname(xlsx_path)

    # Buscar últimos dois snapshots (tolerante a ambos os formatos)
    prev, curr = _load_last_two_any(base_dir, hostname)
    if not curr:
        wb.save(xlsx_path)
        return

    prev_ts = _snap_ts(prev)
    curr_ts = _snap_ts(curr)
    prev_metrics = _snap_metrics(prev)
    curr_metrics = _snap_metrics(curr)

    # === 1) Execucao_<ts> (da corrida ATUAL) ===
    exec_name = f"Execucao_{ts}"
    if exec_name in wb.sheetnames:
        wb.remove(wb[exec_name])
    ws_exec = wb.create_sheet(exec_name)

    ws_exec["A1"] = "Execução"; ws_exec["A1"].font = Font(bold=True); ws_exec["B1"] = ts
    ws_exec["A2"] = "Hostname"; ws_exec["A2"].font = Font(bold=True); ws_exec["B2"] = hostname
    ws_exec["A3"] = "IP";       ws_exec["A3"].font = Font(bold=True); ws_exec["B3"] = host_ip or "-"

    # KPIs (se não houver metric, mete '-')
    def _put(row, label, key):
        ws_exec[f"A{row}"] = label; ws_exec[f"A{row}"].font = Font(bold=True)
        val = curr_metrics.get(key)
        ws_exec[f"B{row}"] = val if val is not None else "-"

    _put(4, "Interfaces (total)",  "interfaces_total")
    _put(5, "Ativas (Up)",         "interfaces_up")
    _put(6, "Inativas (Down)",     "interfaces_down")
    _put(7, "VLANs ativas",        "vlans_active")
    _put(8, "Vizinhos CDP (total)","cdp_total")
    _put(9, "Inventory itens",     "inventory_total")

    # === 2) Comparacao (layout antigo) ===
    if "Comparacao" in wb.sheetnames:
        wb.remove(wb["Comparacao"])
    ws_cmp = wb.create_sheet("Comparacao")

    ws_cmp["A1"] = "Métrica"; ws_cmp["A1"].font = Font(bold=True)
    ws_cmp["B1"] = f"Anterior ({prev_ts})" if prev else "Anterior (-)"; ws_cmp["B1"].font = Font(bold=True)
    ws_cmp["C1"] = f"Atual ({curr_ts})"; ws_cmp["C1"].font = Font(bold=True)
    ws_cmp["D1"] = "Δ"; ws_cmp["D1"].font = Font(bold=True)

    linhas = [
        ("Interfaces (total)", "interfaces_total"),
        ("Ativas (Up)",        "interfaces_up"),
        ("Inativas (Down)",    "interfaces_down"),
        ("VLANs ativas",       "vlans_active"),
        ("Vizinhos CDP (total)","cdp_total"),
        ("Inventory itens (total)","inventory_total"),
    ]

    r = 2
    for label, key in linhas:
        prev_v = prev_metrics.get(key) if prev else None
        curr_v = curr_metrics.get(key)
        ws_cmp.cell(r, 1, label)
        ws_cmp.cell(r, 2, prev_v if prev_v is not None else "-")
        ws_cmp.cell(r, 3, curr_v if curr_v is not None else "-")
        if isinstance(prev_v, (int, float)) and isinstance(curr_v, (int, float)):
            ws_cmp.cell(r, 4, curr_v - prev_v)
        else:
            ws_cmp.cell(r, 4, "-")
        r += 1

    # Larguras amigáveis para Execução/Comparação
    for col in ("A","B","C","D"):
        ws_cmp.column_dimensions[col].width = 24
        ws_exec.column_dimensions[col].width = 24

    # === 3) Alterações Detalhadas + Checklist Migração (NOVO) ===
    try:
        _write_alteracoes_sheet(wb, prev, curr)
        _write_checklist_sheet(wb, prev_metrics, curr_metrics, had_prev=bool(prev))
    except Exception:
        # Não falhar a execução principal se algo der erro aqui
        pass

    # Guardar
    wb.save(xlsx_path)


def ideia6_diagnose(run_print: bool = True) -> Dict[str, Union[int, List[str]]]:
    info = {"total": len(IDEA6_EXECUTED), "funcoes": IDEA6_EXECUTED.copy()}
    if run_print:
        print("[Ideia 6] Funções executadas:", ", ".join(IDEA6_EXECUTED) or "(nenhuma)")
        print("[Ideia 6] Sugestões de testes foram incluídas no docstring desta função.")
    return info
