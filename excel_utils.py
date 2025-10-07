
from __future__ import annotations
from typing import Iterable, List, Tuple, Optional, Any, Dict, Set
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter, coordinate_to_tuple
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.chart import BarChart, PieChart, Reference
import re

# -------- naming & sizing --------
def safe_sheetname(name: str, max_len: int = 31) -> str:
    name = re.sub(r'[:\\/\?*\[\]]', '_', name or 'Sheet')
    return name[:max_len]

def autosize_columns(ws: Worksheet, extra_pad: int = 2, maxw: int = 80) -> None:
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ''
            except Exception:
                val = ''
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + extra_pad, maxw)

# -------- low-level writers --------
def insert_table(ws: Worksheet, start_row: int, headers: List[str], rows: List[List[Any]]) -> tuple[int, int]:
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=start_row, column=j, value=h)
        c.font = Font(bold=True)
    for i, row in enumerate(rows, start=start_row + 1):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.alignment = Alignment(wrap_text=True, vertical="top")
    end_row = start_row + 1 + len(rows)
    end_col = len(headers)
    return end_row, end_col

def write_kv_table(ws: Worksheet, anchor_row: int, anchor_col: int, title: str, data_pairs: List[tuple]) -> tuple[int, int]:
    ws.cell(row=anchor_row, column=anchor_col, value=title).font = Font(bold=True)
    r = anchor_row + 1
    for k, v in data_pairs:
        ws.cell(row=r, column=anchor_col, value=k)
        ws.cell(row=r, column=anchor_col + 1, value=v)
        r += 1
    return r - 1, anchor_col + 1

# -------- charts (A1-based) --------
def _ref(ws: Worksheet, first_cell: str, last_cell: str) -> Reference:
    fr, fc = coordinate_to_tuple(first_cell)
    lr, lc = coordinate_to_tuple(last_cell)
    return Reference(ws, min_col=fc, min_row=fr, max_col=lc, max_row=lr)

def chart_bar(ws: Worksheet, title: str, cats_first: str, cats_last: str, vals_first: str, vals_last: str, anchor: str) -> None:
    fr, fc = coordinate_to_tuple(cats_first)
    lr, lc = coordinate_to_tuple(cats_last)
    if lr < fr:
        return
    chart = BarChart(); chart.title = title
    cats = _ref(ws, cats_first, cats_last)
    vals = _ref(ws, vals_first, vals_last)
    chart.add_data(vals, titles_from_data=False)
    chart.set_categories(cats)
    ws.add_chart(chart, anchor)

def chart_pie(ws: Worksheet, title: str, cats_first: str, cats_last: str, vals_first: str, vals_last: str, anchor: str) -> None:
    fr, fc = coordinate_to_tuple(cats_first)
    lr, lc = coordinate_to_tuple(cats_last)
    if lr < fr:
        return
    chart = PieChart(); chart.title = title
    cats = _ref(ws, cats_first, cats_last)
    vals = _ref(ws, vals_first, vals_last)
    chart.add_data(vals, titles_from_data=False)
    chart.set_categories(cats)
    ws.add_chart(chart, anchor)

# -------- simple CF helpers (headers must exist in row=header_row) --------
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import column_index_from_string

def _last_row(ws: Worksheet, start: int) -> int:
    max_r = ws.max_row or start
    for r in range(max_r, start - 1, -1):
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=r, column=c).value not in (None, ''):
                return r
    return start

def _col_letter_by_header(ws: Worksheet, header_row: int, names: set[str]) -> str | None:
    low = {str(ws.cell(header_row, c).value or '').strip().lower(): c for c in range(1, ws.max_column + 1)}
    for want in names:
        for k, c in low.items():
            if k == want.lower():
                from openpyxl.utils import get_column_letter
                return get_column_letter(c)
    return None

def apply_cf_interfaces(ws: Worksheet, header_row: int = 1) -> None:
    vlan_col = _col_letter_by_header(ws, header_row, {"vlan"})
    status_col = _col_letter_by_header(ws, header_row, {"status"})
    if not vlan_col or not status_col:
        return
    first = header_row + 1
    last = _last_row(ws, first)
    if last < first:
        return

    # Verde na coluna VLAN quando a porta está up/connected e tem VLAN válida (já existia)
    green_fill_vlan = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    formula_up_with_vlan = (
        f'AND(OR(ISNUMBER(SEARCH("connected",${status_col}{first})), ISNUMBER(SEARCH("up",${status_col}{first}))),'
        f'AND(NOT(ISBLANK(${vlan_col}{first})), ${vlan_col}{first}<>"-", ${vlan_col}{first}<>"trunk"))'
    )
    ws.conditional_formatting.add(
        f"${vlan_col}{first}:${vlan_col}{last}",
        FormulaRule(formula=[formula_up_with_vlan], fill=green_fill_vlan)
    )

    # CINZENTO na coluna Status quando "notconnect"
    grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    formula_notconnect = f'ISNUMBER(SEARCH("notconnect",${status_col}{first}))'
    ws.conditional_formatting.add(
        f"${status_col}{first}:${status_col}{last}",
        FormulaRule(formula=[formula_notconnect], fill=grey_fill)
    )

    # NOVO: VERDE na coluna Status quando "connected"
    green_fill_status = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    formula_connected = f'ISNUMBER(SEARCH("connected",${status_col}{first}))'
    ws.conditional_formatting.add(
        f"${status_col}{first}:${status_col}{last}",
        FormulaRule(formula=[formula_connected], fill=green_fill_status)
    )


def apply_cf_vlans(ws: Worksheet, header_row: int = 1) -> None:
    state_col = _col_letter_by_header(ws, header_row, {"status", "state"})
    if not state_col:
        return
    first = header_row + 1
    last = _last_row(ws, first)
    if last < first:
        return
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ws.conditional_formatting.add(f"${state_col}{first}:${state_col}{last}", FormulaRule(formula=[f'EXACT(${state_col}{first},"active")'], fill=green))
    ws.conditional_formatting.add(f"${state_col}{first}:${state_col}{last}", FormulaRule(formula=[f'EXACT(${state_col}{first},"suspend")'], fill=yellow))

def apply_cf_spanning_tree(ws: Worksheet, header_row: int = 1) -> None:
    state_col = _col_letter_by_header(ws, header_row, {"state", "estado"})
    role_col  = _col_letter_by_header(ws, header_row, {"role", "papel"})
    first = header_row + 1
    last  = _last_row(ws, first)

    # Sem coluna 'state' ou sem linhas -> sair silenciosamente
    if not state_col or last < first:
        return

    green  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    red    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    blue   = PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid")

    try:
        ws.conditional_formatting.add(
            f"${state_col}{first}:${state_col}{last}",
            FormulaRule(formula=[f'EXACT(${state_col}{first},"forwarding")'], fill=green)
        )
        ws.conditional_formatting.add(
            f"${state_col}{first}:${state_col}{last}",
            FormulaRule(formula=[f'OR(EXACT(${state_col}{first},"learning"),EXACT(${state_col}{first},"listening"))'], fill=yellow)
        )
        ws.conditional_formatting.add(
            f"${state_col}{first}:${state_col}{last}",
            FormulaRule(formula=[f'EXACT(${state_col}{first},"blocking")'], fill=red)
        )
        if role_col:
            ws.conditional_formatting.add(
                f"${role_col}{first}:${role_col}{last}",
                FormulaRule(formula=[f'EXACT(${role_col}{first},"root")'], fill=blue)
            )
    except Exception:
        # Falha suave para qualquer edge-case de intervalos
        return


def mark_portchannels_and_link(ws: Worksheet, header_row: int = 5, target_sheet_title: str = "show etherchannel summary") -> None:
    from openpyxl.utils import column_index_from_string
    port_col_letter = _col_letter_by_header(ws, header_row, {"port"})
    if not port_col_letter:
        return
    port_col = column_index_from_string(port_col_letter)
    last_col = ws.max_column
    purple_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    first = header_row + 1
    last  = _last_row(ws, first)
    if last < first:
        return
    ether_sheet = safe_sheetname(target_sheet_title)
    for r in range(first, last + 1):
        v = (str(ws.cell(row=r, column=port_col).value).strip().lower()
             if ws.cell(row=r, column=port_col).value is not None else "")
        if v.startswith(("po", "port-channel")):
            for c in range(1, last_col + 1):
                ws.cell(row=r, column=c).fill = purple_fill
            cell = ws.cell(row=r, column=port_col)
            cell.hyperlink = f"#'{ether_sheet}'!A1"
            try:
                cell.font = Font(color="0563C1", underline="single")
            except Exception:
                pass


# -------- Resumo (Ideia 5) --------
def write_resumo_block(ws: Worksheet, title: str, pairs: list[tuple], start_row: int | None = None, start_col: int = 1) -> tuple[int, int]:
    """Escreve um bloco 'Resumo' padronizado (Ideia 5).

    Args:
        ws: Worksheet alvo.
        title: Título do bloco (ex.: 'Resumo – Interfaces').
        pairs: Lista [(chave, valor), ...].
        start_row: Linha inicial; se None, escreve após último conteúdo.
        start_col: Coluna inicial (1=A).
    Returns:
        (last_row, last_col) posicionamento final do bloco.
    """
    # encontrar última linha com conteúdo
    if start_row is None:
        r = ws.max_row or 1
        # avança até encontrar a 1ª linha vazia com alguma folga
        r += 2
    else:
        r = start_row

    ws.cell(row=r, column=start_col, value=title).font = Font(bold=True)
    r += 1
    for k, v in pairs:
        ws.cell(row=r, column=start_col,     value=str(k))
        ws.cell(row=r, column=start_col + 1, value=v)
        r += 1
    autosize_columns(ws)
    return (r - 1, start_col + 1)


#-------------------------------------------------------------------------------
# -------- Execucao sheet helpers --------
from datetime import datetime
from openpyxl.workbook import Workbook

_EXEC_PAT = re.compile(r"^(execu[cç][aã]o)_(\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2})$", re.IGNORECASE)

def _list_execucao_sheets(wb: Workbook):
    """Devolve lista de (ws, name, index, dt or None) para sheets Execução_*."""
    out = []
    for idx, ws in enumerate(wb.worksheets):
        name = ws.title
        m = _EXEC_PAT.match(name.strip())
        dt = None
        if m:
            try:
                dt = datetime.strptime(m.group(2), "%Y-%m-%d_%H-%M-%S")
            except Exception:
                dt = None
            out.append((ws, name, idx, dt))
        elif name.strip().lower() in ("execução", "execucao"):
            out.append((ws, name, idx, None))
    return out

def _move_sheet_to_index(wb: Workbook, ws, target_index: int):
    """Move a sheet para um índice exato (mantém ordem das restantes)."""
    sheets = wb._sheets  # API interna do openpyxl, estável o suficiente para reordenar
    cur = sheets.index(ws)
    if cur == target_index:
        return
    sheets.pop(cur)
    sheets.insert(target_index, ws)


# === VLAN helpers (fonte única, reutilizável) ===
import re as _re

def _idx_contains(headers, candidates):
    if not headers:
        return None
    low = [str(h or "").strip().lower() for h in headers]
    for i, h in enumerate(low):
        for cand in candidates:
            if cand in h:
                return i
    return None

def extract_active_vlans_from_table(headers, rows):
    """
    Devolve set(str) com VLANs ATIVAS a partir de headers/rows de 'show vlan brief'.
    - tolera cabeçalhos: 'vlan', 'vlan id', 'status'/'state'
    - se não houver coluna de estado, assume ativa (templates minimalistas)
    """
    active = set()
    if not headers or not rows:
        return active
    vidx = _idx_contains(headers, ["vlan", "vlan id", "vlan-id"])
    sidx = _idx_contains(headers, ["status", "state"])
    for r in rows:
        vid = (str(r[vidx]).strip() if (vidx is not None and vidx < len(r)) else "")
        st  = (str(r[sidx]).strip().lower() if (sidx is not None and sidx < len(r) and r[sidx]) else "")
        if vid.isdigit() and (("active" in st) or (sidx is None)):
            active.add(vid)
    return active

def extract_active_vlans_from_raw(raw_text):
    """Fallback: procura linhas com 'active' no texto bruto do 'show vlan brief'."""
    active = set()
    if not raw_text:
        return active
    for ln in str(raw_text).splitlines():
        m = _re.match(r"\s*(\d{1,4})\s+.*?\bactive\b", ln, flags=_re.IGNORECASE)
        if m:
            active.add(m.group(1))
    return active
