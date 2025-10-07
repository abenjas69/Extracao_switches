# ADD/CONFIRMAR estes:
from typing import Dict, List, Tuple, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Side, Border   # <- inclui Border
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import coordinate_to_tuple, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
import re



def _find_sheet_by_contains(wb, needle: str):
    needle = needle.lower()
    for ws in wb.worksheets:
        if needle in ws.title.lower():
            return ws
    return None

def _get_table_from_sheet(ws):
    """
    Extrai (headers, rows) de forma robusta:
      1) Se houver Tabela Excel (ws.tables), usa o intervalo da tabela.
      2) Senão, ignora metadados iniciais e deteta a 1ª linha que parece cabeçalho real.
    """
    # 1) Preferir Tabela Excel se existir
    tbl = None
    if getattr(ws, "tables", None):
        tbl = next(iter(ws.tables.values()), None)
    if tbl:
        from openpyxl.utils.cell import range_boundaries
        min_col, min_row, max_col, max_row = range_boundaries(tbl.ref)
        headers = [str(ws.cell(min_row, c).value or "").strip() for c in range(min_col, max_col + 1)]
        rows = []
        for r in range(min_row + 1, max_row + 1):
            row = [ws.cell(r, c).value for c in range(min_col, max_col + 1)]
            # descarta linhas totalmente vazias dentro do ref da Tabela
            if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
                continue
            rows.append(row)
        return headers, rows

    # 2) Fallback: heurística (ignorar metadados e procurar cabeçalho "real")
    header_row_idx = None
    header_tokens = {
        "port", "status", "vlan", "vlan id", "name", "descr", "device id", "platform",
        "local interface", "port id", "uptime", "version", "pid", "speed", "duplex", "type"
    }
    metadata = {"comando:", "hostname:", "timestamp:", "host:", "ip:"}

    for i in range(1, ws.max_row + 1):
        row_vals = [ws.cell(i, c).value for c in range(1, ws.max_column + 1)]
        lrow = [str(v).strip().lower() if v is not None else "" for v in row_vals]
        if not any(lrow):
            continue
        if (lrow[0] in metadata) and sum(1 for v in lrow if v) <= 2:
            # linha de metadados típica, ignora
            continue
        if (sum(1 for v in lrow if v) >= 2) and any(tok in lrow for tok in header_tokens):
            header_row_idx = i
            break

    if header_row_idx is None:
        return [], []

    headers = [str(ws.cell(header_row_idx, c).value or "").strip() for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(header_row_idx + 1, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        # termina à primeira linha totalmente vazia após a tabela
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            break
        rows.append(row)
    return headers, rows


def _index_of_header(headers, *candidates):
    hmap = {h.strip().lower(): i for i, h in enumerate(headers)}
    for cand in candidates:
        idx = hmap.get(cand.lower())
        if idx is not None:
            return idx
    # tentativa soft por contém
    for i, h in enumerate(headers):
        hl = h.lower()
        if any(cand.lower() in hl for cand in candidates):
            return i
    return None

def _parse_interfaces_status(ws_interfaces):
    """
    A partir da folha 'show interfaces status' (ou equivalente),
    conta total / up / down e agrega por VLAN (quando existir coluna VLAN).
    """
    total = up = down = 0
    per_vlan = {}  # vlan_id -> count of ports (qualquer estado)

    headers, rows = _get_table_from_sheet(ws_interfaces)
    if not headers:
        return dict(total=0, up=0, down=0, per_vlan={})

    idx_status = _index_of_header(headers, "status")
    idx_vlan   = _index_of_header(headers, "vlan")
    idx_port   = _index_of_header(headers, "port")  # índice da coluna com o nome da interface


    for row in rows:
        # --- NÃO contar interfaces lógicas (Port-channel/Po, VLAN, Loopback, Tunnel, NVE, Virtual) ---
        if idx_port is not None and idx_port < len(row):
            ifname = (str(row[idx_port]).strip().lower() if row[idx_port] is not None else "")
            if ifname.startswith(("po", "port-channel", "vlan", "lo", "loopback", "tunnel", "nve", "virtual")):
                continue  # salta as lógicas: não entram no 'total', 'up' ou 'down'

        total += 1
        status_val = str(row[idx_status]).strip().lower() if idx_status is not None and idx_status < len(row) else ""
        if status_val in ("connected", "up"):
            up += 1
        else:
            # Cisco muitas vezes usa "notconnect", "err-disabled", "down", etc.
            down += 1

        if idx_vlan is not None and idx_vlan < len(row) and row[idx_vlan] not in (None, "", "trunk"):
            vlan_id = str(row[idx_vlan]).strip()
            per_vlan[vlan_id] = per_vlan.get(vlan_id, 0) + 1

    return dict(total=total, up=up, down=down, per_vlan=per_vlan)

def _is_logical_iface(ifname: str) -> bool:
    """True se 'ifname' for interface lógica (Port-channel/Po, VLAN, Loopback, Tunnel, NVE, Virtual)."""
    if not ifname:
        return False
    n = str(ifname).strip().lower()
    return n.startswith(("po", "port-channel", "vlan", "lo", "loopback", "tunnel", "nve", "virtual"))

def _standardize_pc_name(ifname: str) -> str:
    """Normaliza 'Po1' / 'Port-channel1' / 'Port Channel 1' -> 'Port-channel1'."""
    import re
    s = str(ifname).strip()
    m = re.search(r'(?i)\bpo(?:rt-?channel)?\s*([0-9]+)\b', s)
    return f"Port-channel{m.group(1)}" if m else s

def parser_portchannels_from_interfaces_sheet(ws):
    """
    Lê a folha 'show interfaces status' e devolve:
      { 'Port-channel1': {'members': ['Gi1/0/1','Gi1/0/2']}, ... }
    Heurística:
      1) deteta Port-channels pela coluna 'Port';
      2) associa membros físicos pela coluna 'Name/Description' quando menciona 'PoX'/'Port-channelX'.
    """
    import re
    headers, rows = _get_table_from_sheet(ws)
    if not headers or not rows:
        return {}

    # índices tolerantes
    def _idx_of(*cands):
        return _index_of_header(headers, *cands)

    idx_port = _idx_of("port")
    idx_name = _idx_of("name", "descr", "description")

    pcs = {}

    # (1) detetar Port-channels pela própria coluna 'Port'
    if idx_port is not None:
        for r in rows:
            port = str(r[idx_port]).strip() if idx_port < len(r) and r[idx_port] is not None else ""
            if port and _is_logical_iface(port):
                norm = _standardize_pc_name(port)
                if norm.lower().startswith("port-channel"):
                    pcs.setdefault(norm, {"members": []})

    # (2) inferir membros físicos via Name/Description
    if idx_port is not None and idx_name is not None:
        for r in rows:
            phys = str(r[idx_port]).strip() if idx_port < len(r) and r[idx_port] is not None else ""
            if not phys or _is_logical_iface(phys):
                continue
            name_desc = str(r[idx_name]).strip() if idx_name < len(r) and r[idx_name] is not None else ""
            if not name_desc:
                continue
            m = re.search(r'(?i)\bpo(?:rt-?channel)?\s*([0-9]+)\b', name_desc)
            if m:
                pc = f"Port-channel{m.group(1)}"
                pcs.setdefault(pc, {"members": []})
                if phys not in pcs[pc]["members"]:
                    pcs[pc]["members"].append(phys)
    return pcs

def _parse_vlans(ws_vlans):
    """
    A partir da folha 'show vlan brief' (ou similar),
    apanha (VLAN, Nome, Estado).
    """
    out = []
    headers, rows = _get_table_from_sheet(ws_vlans)
    if not headers:
        return out

    idx_vlan  = _index_of_header(headers, "vlan", "vlan id")
    idx_name  = _index_of_header(headers, "name")
    idx_state = _index_of_header(headers, "status", "state")

    for row in rows:
        vlan_id = str(row[idx_vlan]).strip() if idx_vlan is not None and idx_vlan < len(row) else ""
        if not vlan_id.isdigit():
            # ignora headings intermédios ou linhas lixo
            continue
        name = str(row[idx_name]).strip() if idx_name is not None and idx_name < len(row) and row[idx_name] else ""
        state = str(row[idx_state]).strip().lower() if idx_state is not None and idx_state < len(row) and row[idx_state] else ""
        out.append((vlan_id, name, state))
    return out

def _parse_show_version(ws_version):
    """
    Extrai 'IOS version' e 'uptime' de forma robusta:
      - Primeiro, tenta 'Resumo_Version' (B3/B4).
      - Depois, tenta Tabela da folha 'show version' (colunas 'version' e 'uptime').
      - Por fim, fallback por regex varrendo todas as células.
    """
    # 1) Usar 'Resumo_Version' se existir
    wb = ws_version.parent
    if "Resumo_Version" in wb.sheetnames:
        rv = wb["Resumo_Version"]
        ver = (str(rv["B3"].value).strip() if rv["B3"].value is not None else "")
        up  = (str(rv["B4"].value).strip() if rv["B4"].value is not None else "")
        if ver or up:
            return (ver or "Desconhecida", up or "Desconhecido")

    # 2) Usar Tabela da própria folha (quando TextFSM criou estrutura tabular)
    tbl = None
    if getattr(ws_version, "tables", None):
        tbl = next(iter(ws_version.tables.values()), None)
    if tbl:
        from openpyxl.utils.cell import range_boundaries
        min_col, min_row, max_col, max_row = range_boundaries(tbl.ref)
        headers = [str(ws_version.cell(min_row, c).value or "").strip().lower()
                   for c in range(min_col, max_col + 1)]
        hmap = {h: i for i, h in enumerate(headers)}
        row1 = [ws_version.cell(min_row + 1, c).value for c in range(min_col, max_col + 1)]
        ver = (str(row1[hmap["version"]]).strip() if "version" in hmap and row1[hmap["version"]] is not None else "")
        up  = (str(row1[hmap["uptime"]]).strip()  if "uptime"  in hmap and row1[hmap["uptime"]]  is not None else "")
        if ver or up:
            return (ver or "Desconhecida", up or "Desconhecido")

    # 3) Fallback: regex em todas as células (não apenas coluna A)
    text_lines = []
    for r in range(1, ws_version.max_row + 1):
        for c in range(1, ws_version.max_column + 1):
            val = ws_version.cell(r, c).value
            if val is None:
                continue
            text_lines.append(str(val))
    whole = "\n".join(text_lines)

    m = re.search(r"Version\s+([\w.\(\)-]+)", whole, re.IGNORECASE)
    ios_version = m.group(1) if m else "Desconhecida"

    mu = re.search(r"uptime\s+is\s+([^\n]+)", whole, re.IGNORECASE)
    uptime = mu.group(1).strip() if mu else "Desconhecido"

    return ios_version, uptime


def parser_portchannels_from_etherchannel_sheet(ws):
    """
    Lê a folha 'show etherchannel summary' e devolve:
      { 'Port-channel1': {'members': ['Gi1/0/1','Gi1/0/2']}, ... }

    Suporta variações de cabeçalhos dos templates (TextFSM/Netmiko):
      - Port-Channel: 'port-channel', 'port_channel', 'portchannel', 'po', 'bundle', ou 'group/channel_group' (número)
      - Membros:      'member_interface', 'members', 'ports', 'interfaces', 'port_list'
    """
    import re, ast
    headers, rows = _get_table_from_sheet(ws)
    if not headers or not rows:
        return {}

    hmap = {str(h).strip().lower(): i for i, h in enumerate(headers)}
    # localizar colunas (robusto a variações)
    idx_pc = next((i for k, i in hmap.items() if any(x in k for x in (
        "port-channel", "port_channel", "portchannel", "po", "bundle"))), None)
    idx_grp = next((i for k, i in hmap.items() if "group" in k), None)
    idx_mem = next((i for k, i in hmap.items() if any(x in k for x in (
        "member", "members", "ports", "interfaces", "port_list"))), None)

    def _std_pc_name(raw, grp):
        if raw:
            return _standardize_pc_name(str(raw))
        if grp:
            g = re.sub(r"\D+", "", str(grp))
            return f"Port-channel{g}" if g else str(grp)
        return "Port-channel?"

    def _split_members(val):
        """
        Aceita:
          - lista Python em string: "['Gi1/0/1(P)','Gi1/0/2(P)']"
          - string com espaços/vírgulas: "Gi1/0/1(P) Gi1/0/2(P)"
        Limpa sufixos como '(P)', '(I)', etc.
        """
        if val is None:
            return []
        s = str(val).strip()
        out = []
        # tentar lista Python
        try:
            parsed = ast.literal_eval(s)
            if isinstance(parsed, (list, tuple)):
                cand = [str(x) for x in parsed]
            else:
                cand = re.split(r"[,\s]+", s)
        except Exception:
            cand = re.split(r"[,\s]+", s)

        for t in cand:
            t = t.strip().strip(",;")
            if not t:
                continue
            t = re.sub(r"\(.*?\)$", "", t)   # remove (P), (I), (w), ...
            if not _is_logical_iface(t):     # garantir que não é Po*/Vlan/Loopback
                out.append(t)
        return out

    pcs = {}
    for r in rows:
        pc_name = _std_pc_name(
            r[idx_pc] if (idx_pc is not None and idx_pc < len(r)) else "",
            r[idx_grp] if (idx_grp is not None and idx_grp < len(r)) else ""
        )
        members = _split_members(r[idx_mem] if (idx_mem is not None and idx_mem < len(r)) else "")
        if not pc_name:
            continue
        entry = pcs.setdefault(pc_name, {"members": []})
        for m in members:
            if m and m not in entry["members"]:
                entry["members"].append(m)
    return pcs


# [DEBUG-ETHERCHANNEL] Parser robusto para 'show etherchannel summary' (IOS/NX-OS/NX)
def parse_etherchannel_summary_from_text(raw_text: str) -> Tuple[List[str], List[List[str]]]:
    """
    Retorna (headers, rows) com schema estável:
      ['Group','Port-Channel','Protocol','Status','Flags','Member Ports']

    - Consolida linhas de continuação dos ports numa única célula (Member Ports).
    - Deriva 'Status' a partir das flags: 'U' -> Up, 'D' -> Down, caso contrário 'Unknown'.
    - Tolerante a variações de espaço, tabs, CRLF e banners.
    """
    if not raw_text:
        return ([], [])

    import re
    s = raw_text.replace("\r\n", "\n").replace("\r", "\n").replace("\t", " ")
    # remover caracteres não imprimíveis (menos \n)
    s = "".join(ch for ch in s if (ch == "\n") or ch.isprintable())
    # colapsar espaços múltiplos
    s = re.sub(r"[ ]{2,}", " ", s)

    lines = s.split("\n")

    # Encontrar início da tabela (linha com Group/Port-Channel/Protocol)
    header_idx = None
    header_re = re.compile(r"\bGroup\b.*\bPort-?Channel\b.*\bProtocol\b", re.IGNORECASE)
    for i, ln in enumerate(lines):
        if header_re.search(ln):
            header_idx = i
            break
    start = (header_idx + 1) if header_idx is not None else 0

    # Linha de dados típica: "<grp> PoX(<flags>) <protocol> <ports...>"
    data_line_re = re.compile(
        r"^\s*(?P<group>\d+)\s+(?P<po>[A-Za-z]+[0-9]+)\((?P<flags>[^)]+)\)\s+(?P<protocol>\S+)\s*(?P<ports>.*)$"
    )

    entries = []
    cur = None
    for ln in lines[start:]:
        if not ln.strip():
            continue
        m = data_line_re.match(ln)
        if m:
            if cur:
                entries.append(cur)
            cur = {
                "group": m.group("group"),
                "port_channel": m.group("po").strip(),
                "flags": m.group("flags").strip(),
                "protocol": m.group("protocol").strip(),
                "member_ports": (m.group("ports") or "").strip(),
            }
        else:
            # Continuação das portas (linhas que não começam por dígito)
            if cur and not re.match(r"^\s*\d+\s+", ln):
                extra = ln.strip()
                if extra:
                    cur["member_ports"] = (cur["member_ports"] + " " + extra).strip()

    if cur:
        entries.append(cur)

    # Normalização final de member_ports e derivação de status
    rows = []
    for e in entries:
        ports = e["member_ports"]
        # separar por vírgulas ou espaços, limpar vazios e voltar a juntar com ", "
        toks = re.split(r"[,\s]+", ports) if ports else []
        toks = [t for t in toks if t]
        member_ports = ", ".join(toks)

        flags = (e["flags"] or "").upper()
        status = "Up" if "U" in flags else ("Down" if "D" in flags else "Unknown")

        rows.append([
            int(e["group"]),
            e["port_channel"],
            e["protocol"],
            status,
            e["flags"],
            member_ports
        ])

    headers = ["Group","Port-Channel","Protocol","Status","Flags","Member Ports"]
    return (headers, rows)

def parser_show_spanning_tree_from_text(raw_text: str) -> Tuple[List[str], List[List[str]]]:
    """
    Parser interno para 'show spanning-tree' (PVST/RSTP).
    Gera uma tabela com: vlan, interface, role, state, cost, port_id, port_type.
    Funciona com o formato típico IOS:
      VLAN0001 / VLAN 1 ... + tabela de portas "Interface  Role Sts Cost  Prio.Nbr Type"
    """
    import re

    if not raw_text:
        return ([], [])

    t = raw_text.replace("\r\n", "\n")
    headers = ["vlan", "interface", "role", "state", "cost", "port_id", "port_type"]
    rows: List[List[str]] = []

    # Encontrar blocos por VLAN/instância
    # ex.: "VLAN0001" ou "VLAN 1"
    vlan = None
    vlan_hdr = re.compile(r"^\s*VLAN\s*0*(\d+)\b", re.IGNORECASE)

    # Linha de porta típica:
    # Gi1/0/1   Desg FWD 19 128.1 P2p
    port_re = re.compile(
        r'^(?P<intf>\S+)\s+(?P<role>\w+)\s+(?P<state>\w+)\s+(?P<cost>\d+)\s+(?P<portid>[\d\.]+)\s+(?P<ptype>.+?)\s*$',
        re.IGNORECASE
    )

    for line in t.split("\n"):
        m_vlan = vlan_hdr.search(line)
        if m_vlan:
            vlan = m_vlan.group(1)
            continue

        m_port = port_re.match(line.strip())
        if m_port and vlan:
            rows.append([
                vlan,
                m_port.group("intf"),
                m_port.group("role").lower(),
                m_port.group("state").lower(),
                m_port.group("cost"),
                m_port.group("portid"),
                m_port.group("ptype"),
            ])

    return (headers, rows if rows else [])



def build_dashboard(wb):
    # 1) Encontrar as folhas relevantes
    ws_int = _find_sheet_by_contains(wb, "show interfaces status")
    ws_vl  = _find_sheet_by_contains(wb, "show vlan brief")
    ws_ver = _find_sheet_by_contains(wb, "show version")
    ws_eth = _find_sheet_by_contains(wb, "etherchannel")


    # 2) Extrair métricas
    iface_stats = _parse_interfaces_status(ws_int) if ws_int else dict(total=0, up=0, down=0, per_vlan={})
    vlans = _parse_vlans(ws_vl) if ws_vl else []
    ios_version, uptime = _parse_show_version(ws_ver) if ws_ver else ("Desconhecida", "Desconhecido")
    pc_map = (parser_portchannels_from_etherchannel_sheet(ws_eth) if ws_eth
            else (parser_portchannels_from_interfaces_sheet(ws_int) if ws_int else {}))



    # 3) Criar / limpar Dashboard
    if "Dashboard" in wb.sheetnames:
        ws_dash = wb["Dashboard"]
        wb.remove(ws_dash)
    ws_dash = wb.create_sheet("Dashboard", 0)  # 1.ª aba

    # 4) Cabeçalho e KPIs (à esquerda)
    ws_dash["A1"] = "Dashboard – Resumo do Switch"
    ws_dash["A1"].font = Font(size=16, bold=True)
    ws_dash.merge_cells("A1:D1")

    ws_dash["A3"] = "Total de Interfaces"
    ws_dash["B3"] = iface_stats["total"]
    ws_dash["A4"] = "Ativas (Up)"
    ws_dash["B4"] = iface_stats["up"]
    ws_dash["A5"] = "Inativas (Down/NotConnect)"
    ws_dash["B5"] = iface_stats["down"]

    ws_dash["A7"] = "IOS Version"
    ws_dash["B7"] = ios_version
    ws_dash["A8"] = "Uptime"
    ws_dash["B8"] = uptime

    for cell in ("A3","A4","A5","A7","A8"):
        ws_dash[cell].font = Font(bold=True)    

    # --- Tabela "Port-Channels" à direita dos KPIs ---
    start_col = "D"   # ao lado dos KPIs (A..B ocupados); ajusta se precisares
    start_row = 3

    ws_dash[f"{start_col}{start_row}"] = "Port-Channel"
    ws_dash[f"{chr(ord(start_col)+1)}{start_row}"] = "Membros"
    ws_dash[f"{start_col}{start_row}"].font = Font(bold=True)
    ws_dash[f"{chr(ord(start_col)+1)}{start_row}"].font = Font(bold=True)

    r = start_row + 1
    if pc_map:
        for pc_name, data in sorted(pc_map.items(), key=lambda x: x[0]):
            ws_dash[f"{start_col}{r}"] = pc_name
            members = ", ".join(data.get("members", [])) if data.get("members") else "(sem associação encontrada)"
            ws_dash[f"{chr(ord(start_col)+1)}{r}"] = members
            r += 1
    else:
        ws_dash[f"{start_col}{r}"] = "(nenhum Port-Channel detetado)"
        ws_dash[f"{chr(ord(start_col)+1)}{r}"] = "-"
        r += 1

    # borda simples na tabela (usa 'Side' que já importaste)
    try:
        thin = Side(style="thin")
        for rr in range(start_row, r):
            for cc in range(ord(start_col), ord(start_col)+2):
                ws_dash[f"{chr(cc)}{rr}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
    except Exception:
        pass


    # === (NOVO) KPIs de CDP / Neighbors ===
    ws_cdp = _find_sheet_by_contains(wb, "cdp")
    neighbors_count = 0
    top_platforms = []

    if ws_cdp:
        cdp_headers, cdp_rows = _get_table_from_sheet(ws_cdp)
        if cdp_headers:
            neighbors_count = len(cdp_rows)

            # índice de colunas de forma tolerante a variações
            def _idx_any(headers, *cands):
                for i, h in enumerate(headers):
                    if h is None:
                        continue
                    hl = str(h).strip().lower()
                    if any(c in hl for c in cands):
                        return i
                return None

            idx_platform = _idx_any(cdp_headers, "platform")

            from collections import Counter
            plats = Counter()
            if idx_platform is not None:
                for r in cdp_rows:
                    val = r[idx_platform] if idx_platform < len(r) else ""
                    plat = str(val).split()[0] if val else "Desconhecido"
                    plats[plat] += 1
            top_platforms = plats.most_common(5)

    # KPI principal de CDP junto aos restantes
    ws_dash["A9"] = "Vizinhos (CDP)"
    ws_dash["B9"] = neighbors_count
    ws_dash["A9"].font = Font(bold=True)

    # 5) SECÇÃO VLANs — Nome e Estado (tabela à direita F–H)
    ws_dash.merge_cells("F2:H2")
    ws_dash["F2"] = "VLANs — Nome e Estado"
    ws_dash["F2"].font = Font(bold=True)
    ws_dash["F2"].alignment = Alignment(horizontal="center")

    ws_dash["F3"] = "VLAN"
    ws_dash["G3"] = "Nome"
    ws_dash["H3"] = "Estado"
    for c in ("F3","G3","H3"):
        ws_dash[c].font = Font(bold=True)
        ws_dash[c].alignment = Alignment(horizontal="center")

    row = 4
    for vlan_id, name, state in vlans:
        ws_dash[f"F{row}"] = vlan_id
        ws_dash[f"G{row}"] = name
        ws_dash[f"H{row}"] = state
        row += 1

    thin = Side(style="thin")
    rng = f"F3:H{max(3, row-1)}"
    for r in ws_dash[rng]:
        for cell in r:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 6) Distribuição de Estados (tabela pequena para o Pie)
    ws_dash["A11"] = "Distribuição de Estados"
    ws_dash["A11"].font = Font(bold=True)
    ws_dash["A12"] = "Up"
    ws_dash["A13"] = "Down"
    ws_dash["B12"] = iface_stats["up"]
    ws_dash["B13"] = iface_stats["down"]

    # 7) Gráfico Pizza: Interfaces Up vs Down (à direita, topo)
    pie = PieChart()
    pie.title = "Interfaces Up vs Down"
    labels = Reference(ws_dash, min_col=1, min_row=12, max_row=13)
    data   = Reference(ws_dash, min_col=2, min_row=12, max_row=13)
    pie.add_data(data, titles_from_data=False)
    pie.set_categories(labels)
    ws_dash.add_chart(pie, "J3")

    # 8) Mini-tabela: Portas por VLAN (à esquerda, por baixo dos KPIs)
    if iface_stats["per_vlan"]:
        ws_dash["A16"] = "Portas por VLAN"
        ws_dash["A16"].font = Font(bold=True)
        base = 17
        ws_dash.cell(row=base, column=1, value="VLAN").font = Font(bold=True)
        ws_dash.cell(row=base, column=2, value="Nº Portas").font = Font(bold=True)
        for i, (vlan_id, count) in enumerate(
            sorted(
                iface_stats["per_vlan"].items(),
                key=lambda kv: int(re.sub(r'[^0-9]', '', kv[0]) or 0)
            )
        ):
            ws_dash.cell(row=base + 1 + i, column=1, value=vlan_id)
            ws_dash.cell(row=base + 1 + i, column=2, value=count)

        last_row = base + 1 + len(iface_stats["per_vlan"])

        # 9) Gráfico Barras: Portas por VLAN (à direita, abaixo do Pie)
        bar = BarChart()
        bar.title = "Portas por VLAN (vista de Interfaces)"
        data = Reference(ws_dash, min_col=2, min_row=base, max_row=last_row)
        cats = Reference(ws_dash, min_col=1, min_row=base+1, max_row=last_row)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        bar.y_axis.title = "N.º de portas"
        bar.x_axis.title = "VLAN"
        ws_dash.add_chart(bar, "J18")

    # 10) (NOVO) Top Plataformas CDP + gráfico posicionado em baixo do gráfico de VLANs
    row_base = 22  # dá um espaço visual depois de "Portas por VLAN"
    if top_platforms:
        ws_dash.cell(row=row_base,   column=1, value="Plataformas CDP").font = Font(bold=True)
        ws_dash.cell(row=row_base+1, column=1, value="Plataforma").font = Font(bold=True)
        ws_dash.cell(row=row_base+1, column=2, value="Quantidade").font = Font(bold=True)
        for i, (plat, n) in enumerate(top_platforms, start=0):
            ws_dash.cell(row=row_base+2+i, column=1, value=plat)
            ws_dash.cell(row=row_base+2+i, column=2, value=n)

        # Gráfico de barras para Top Plataformas (ABAIXO de J18 -> J35)
        bar2 = BarChart()
        bar2.title = "CDP por Plataforma (Top)"
        data2 = Reference(ws_dash, min_col=2, min_row=row_base+1, max_row=row_base+1+len(top_platforms))
        cats2 = Reference(ws_dash, min_col=1, min_row=row_base+2, max_row=row_base+1+len(top_platforms))
        bar2.add_data(data2, titles_from_data=True)
        bar2.set_categories(cats2)
        ws_dash.add_chart(bar2, "J35")  # NÃO tapa J18

    # 11) Ajustes visuais gerais
    ws_dash.column_dimensions["A"].width = 26
    ws_dash.column_dimensions["B"].width = 38
    ws_dash.column_dimensions["F"].width = 8
    ws_dash.column_dimensions["G"].width = 22
    ws_dash.column_dimensions["H"].width = 14
    ws_dash.column_dimensions["D"].width = 18
    ws_dash.column_dimensions["E"].width = 50

    ws_dash.freeze_panes = "A3"

    # Centralizar título
    ws_dash["A1"].alignment = Alignment(horizontal="center")

    # Garantir que fica 1.ª aba
    wb.move_sheet(ws_dash, offset=-wb.index(ws_dash))

