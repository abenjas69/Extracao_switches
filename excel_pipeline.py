
from __future__ import annotations
import os, re, logging
from typing import List, Tuple, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from .excel_utils import safe_sheetname, autosize_columns, insert_table, write_kv_table, chart_bar, chart_pie, write_resumo_block
from .excel_utils import apply_cf_interfaces, apply_cf_vlans, apply_cf_spanning_tree, mark_portchannels_and_link
from .excel_utils import   extract_active_vlans_from_table, extract_active_vlans_from_raw
from .textfsm_utils import parse_with_textfsm, _dictlist_to_table
from .parsers import parser_show_spanning_tree_from_text, parse_etherchannel_summary_from_text, parse_show_interfaces_trunk
from .excel_dashboard import build_dashboard

log = logging.getLogger("clean_switch.excel_pipeline")

def _po_members_from_ether_rows(headers: list[str], rows: list[list[Any]]):
    """Extrai dict {po: [members...]} e lista [(po, 'm1, m2, ...')] a partir da tabela EtherChannel.

    Espera schema: ['Group','Port-Channel','Protocol','Status','Flags','Member Ports']

    """
    if not headers or not rows:
        return {}, []
    low = [h.lower() for h in headers]
    try:
        po_idx = low.index('port-channel')
        mem_idx = low.index('member ports')
    except ValueError:
        return {}, []
    mapping = {}
    listing = []
    for r in rows:
        po = str(r[po_idx]).strip() if po_idx < len(r) else ''
        mem = str(r[mem_idx]).strip() if mem_idx < len(r) else ''
        if not po:
            continue
        members = [t for t in re.split(r'[\s,]+', mem) if t]
        mapping[po] = members
        listing.append((po, ', '.join(members)))
    return mapping, listing


def metrics_from_collected(collected) -> dict:
    def _norm(s): return (s or "").strip().lower()
    m = {
        "interfaces_total": 0,
        "interfaces_up": 0,
        "interfaces_down": 0,
        "vlans_active": 0,
        "cdp_total": 0,
        "inventory_total": 0,
    }
    for cmd, raw, headers, rows in collected:
        cl = _norm(cmd)
        if "show interfaces status" in cl and headers:
            low = [ _norm(x) for x in headers ]
            p_idx = low.index("port") if "port" in low else None
            s_idx = low.index("status") if "status" in low else None
            for r in rows:
                if p_idx is not None and p_idx < len(r):
                    port = (str(r[p_idx]).strip().lower() if r[p_idx] is not None else "")
                    if port.startswith(("po","port-channel","vlan","lo","loopback","tunnel","nve","virtual")):
                        continue
                m["interfaces_total"] += 1
                s = (str(r[s_idx]).strip().lower() if (s_idx is not None and s_idx < len(r) and r[s_idx] is not None) else "")
                if s in ("connected","up"):
                    m["interfaces_up"] += 1
                else:
                    m["interfaces_down"] += 1
        elif "show vlan brief" in cl:
            vset = set()
            if headers and rows:
                vset = extract_active_vlans_from_table(headers, rows)
            if not vset and raw:
                vset = extract_active_vlans_from_raw(raw)
            m["vlans_active"] = len(vset)
        elif "cdp neighbors" in cl:
            m["cdp_total"] = max(m["cdp_total"], len(rows or []))
        elif "show inventory" in cl:
            m["inventory_total"] = max(m["inventory_total"], len(rows or []))
    return m

def _normalize_headers(cmd_lc: str, hdrs: list[str]) -> list[str]:
    if not hdrs:
        return hdrs
    def pick(*cands):
        low = [h.lower() for h in hdrs]
        for c in cands:
            if c.lower() in low:
                return hdrs[low.index(c.lower())]
        return None
    if "show interfaces status" in cmd_lc:
        wanted = ["Port", "Name", "Status", "Vlan", "Duplex", "Speed", "Type"]
        return [pick(w) or w for w in wanted]
    if "show vlan brief" in cmd_lc:
        wanted = ["VLAN", "Name", "Status", "Interfaces"]
        return [pick(w) or w for w in wanted]
    if "show inventory" in cmd_lc:
        wanted = ["Name", "Descr", "PID", "VID", "SN"]
        return [pick(w) or w for w in wanted]
    if "show version" in cmd_lc:
        return hdrs   # não mexer — manter os nomes vindos do TextFSM
    if "show spanning-tree" in cmd_lc:
        mapping = {"vlan":"Vlan","interface":"Interface","role":"Role","state":"State","cost":"Cost","port_id":"Port ID","port_type":"Type"}
        return [mapping.get(h.lower(), h) for h in hdrs]
    return [h[:1].upper() + h[1:] if isinstance(h, str) else h for h in hdrs]

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

def create_excel(hostname: str, ts: str, collected, xlsx_path: str) -> None:
    wb = Workbook()

    # Sheet "Resumo" minimal
    ws_resume = wb.active
    ws_resume.title = "Resumo"
    ws_resume["A1"] = "Hostname";  ws_resume["B1"] = hostname
    ws_resume["A2"] = "Data/Hora"; ws_resume["B2"] = ts
    ws_resume["A3"] = "Comandos";  ws_resume["B3"] = ", ".join([c[0] for c in collected])
    autosize_columns(ws_resume)

    mono = Font(name="Consolas")
    wrap = Alignment(wrap_text=True, vertical="top")

    ether_headers, ether_rows = [], []

    for (cmd, raw, headers, rows) in collected:
        base = safe_sheetname(cmd)
        title = base
        sfx = 2
        while title in wb.sheetnames:
            title = safe_sheetname(f"{base}_{sfx}")
            sfx += 1

        ws = wb.create_sheet(title=title)
        ws["A1"] = "Comando:";   ws["B1"] = cmd
        ws["A2"] = "Hostname:";  ws["B2"] = hostname
        ws["A3"] = "Timestamp:"; ws["B3"] = ts
        ws["A1"].font = ws["A2"].font = ws["A3"].font = Font(bold=True)

        start_row = 5
        cmd_lc = cmd.strip().lower()

        # Se headers/rows vierem vazios, tentar parsers internos (fallback)
        if (not headers) or (not rows):
            h, r = ([], [])
            if "show spanning-tree" in cmd_lc:
                from .parsers import parser_show_spanning_tree_from_text
                h, r = parser_show_spanning_tree_from_text(raw or "")
            elif ("show etherchannel" in cmd_lc) and ("summary" in cmd_lc):
                from .parsers import parse_etherchannel_summary_from_text
                h, r = parse_etherchannel_summary_from_text(raw or "")
            elif "show interfaces trunk" in cmd_lc:
                from .parsers import parse_show_interfaces_trunk
                h, r = parse_show_interfaces_trunk(raw or "")
            if h and r:
                headers, rows = h, r

        if headers and rows:
            # Normalização de cabeçalhos (NOTA: show version não renomear)
            headers = _normalize_headers(cmd_lc, list(headers))

            data_start_row = start_row
            end_row, end_col = insert_table(ws, data_start_row, headers, rows)
            autosize_columns(ws)

            # Converter em Excel Table (filtros/ordenar)
            ref = f"A{data_start_row}:{get_column_letter(end_col)}{end_row}"
            # Nome único e “safe” para a tabela
            import re
            base_tbl = re.sub(r'[^A-Za-z0-9_]', '_', ws.title) or "Table"
            existing = {t.displayName for t in getattr(ws, "_tables", [])}
            name = base_tbl if base_tbl not in existing else f"{base_tbl}_2"
            i = 2
            while name in existing:
                i += 1
                name = f"{base_tbl}_{i}"

            tbl = Table(displayName=name, ref=ref)
            tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)
            ws.add_table(tbl)

            # Qualidade de vida
            ws.freeze_panes = f"A{data_start_row+1}"

            # Guardar EtherChannel p/ resumos/links
            if "show etherchannel summary" in cmd_lc:
                ether_headers, ether_rows = headers, rows

            # CF por folha
            if "show interfaces status" in cmd_lc:
                apply_cf_interfaces(ws, header_row=data_start_row)
                try:
                    mark_portchannels_and_link(ws, header_row=data_start_row, target_sheet_title="show etherchannel summary")
                except Exception:
                    pass
            elif "show vlan brief" in cmd_lc:
                apply_cf_vlans(ws, header_row=data_start_row)
            elif "show spanning-tree" in cmd_lc:
                # Só aplicar CF se existir a coluna "State"
                if any(str(h).strip().lower() == "state" for h in headers):
                    apply_cf_spanning_tree(ws, header_row=data_start_row)

            # --------- Resumos (Ideia 5) + gráfico rápido por comando ----------
            low = [str(h).lower() for h in headers]

            if "show interfaces status" in cmd_lc:
                sidx = low.index("status") if "status" in low else None
                pidx = low.index("port") if "port" in low else None
                up = down = total = 0
                for r in rows:
                    # ignorar interfaces lógicas (Po/VLAN/Loopback/etc.)
                    if pidx is not None and pidx < len(r):
                        port = str(r[pidx] or "").strip().lower()
                        if port.startswith(("po", "port-channel", "vlan", "lo", "loopback", "tunnel", "nve", "virtual")):
                            continue
                    total += 1
                    st = (str(r[sidx]).strip().lower() if (sidx is not None and sidx < len(r) and r[sidx]) else "")
                    if st in ("connected", "up"):
                        up += 1
                    else:
                        down += 1

                last_r, _ = write_resumo_block(
                    ws, "Resumo – Interfaces",
                    [("TOTAL", total), ("UP", up), ("DOWN", down)],
                    start_row=end_row + 2
                )

                # Port-Channels e membros (se existir EtherChannel)
                if ether_headers and ether_rows:
                    _map, listing = _po_members_from_ether_rows(ether_headers, ether_rows)
                    if listing:
                        ws.cell(row=last_r + 2, column=1, value="Port-Channels e Membros").font = Font(bold=True)
                        rr = last_r + 3
                        for po, members in listing:
                            ws.cell(row=rr, column=1, value=po)
                            ws.cell(row=rr, column=2, value=members)
                            rr += 1
                        autosize_columns(ws)

                # Gráfico “Portas por Estado” (pequeno)
                kv_row = end_row + 2
                chart_bar(
                    ws, "Portas por Estado",
                    f"A{kv_row+1}", f"A{kv_row+2}",
                    f"B{kv_row+1}", f"B{kv_row+2}",
                    "H5"
                )

            elif "show vlan brief" in cmd_lc:
                vidx = low.index("vlan") if "vlan" in low else None
                sidx = low.index("status") if "status" in low else None
                total_vlans = 0
                active = 0
                for r in rows:
                    vid = str(r[vidx]).strip() if (vidx is not None and vidx < len(r)) else ""
                    st  = str(r[sidx]).strip().lower() if (sidx is not None and sidx < len(r) and r[sidx]) else ""
                    if vid.isdigit():
                        total_vlans += 1
                        if st == "active":
                            active += 1
                write_resumo_block(
                    ws, "Resumo – VLANs",
                    [("TOTAL", total_vlans), ("ATIVAS", active)],
                    start_row=end_row + 2
                )
            
            elif "show interfaces trunk" in cmd_lc:
                low = [str(h).strip().lower() for h in headers]
                pidx = low.index("port")            if "port" in low else None
                midx = low.index("mode")            if "mode" in low else None
                sidx = low.index("status")          if "status" in low else None
                nidx = low.index("native_vlan")     if "native_vlan" in low else None

                total_trunks = 0
                trunks_up = 0
                native_not_1 = 0
                natives_list = []  # [(port,native)]

                for r in rows:
                    total_trunks += 1
                    st = str(r[sidx]).strip().lower() if (sidx is not None and sidx < len(r) and r[sidx]) else ""
                    if st in ("up", "connected", "trunking"):
                        trunks_up += 1
                    native = str(r[nidx]).strip() if (nidx is not None and nidx < len(r)) else ""
                    prt = str(r[pidx]).strip() if (pidx is not None and pidx < len(r)) else ""
                    if native and native.isdigit() and native != "1":
                        native_not_1 += 1
                        natives_list.append((prt, native))

                last_r, _ = write_resumo_block(
                    ws, "Resumo – Trunks",
                    [("TOTAL", total_trunks), ("UP", trunks_up), ("NATIVE ≠ 1", native_not_1)],
                    start_row=end_row + 2
                )

                # Lista rápida das natives != 1, se existir
                if natives_list:
                    ws.cell(row=last_r + 2, column=1, value="Portas com Native VLAN ≠ 1").font = Font(bold=True)
                    rr = last_r + 3
                    for prt, nv in natives_list:
                        ws.cell(row=rr, column=1, value=prt)
                        ws.cell(row=rr, column=2, value=nv)
                        rr += 1
                    autosize_columns(ws)


            elif "show inventory" in cmd_lc:
                write_resumo_block(ws, "Resumo – Inventory", [("Itens", len(rows))], start_row=end_row + 2)

            elif "cdp neighbors" in cmd_lc:
                write_resumo_block(ws, "Resumo – CDP", [("Vizinhos", len(rows))], start_row=end_row + 2)

            elif "show version" in cmd_lc:
                v = rows[0][low.index("version")] if ("version" in low and rows) else ""
                u = rows[0][low.index("uptime")]  if ("uptime" in low and rows)  else ""
                write_resumo_block(ws, "Resumo – Version", [("Version", v), ("Uptime", u)], start_row=end_row + 2)

            elif "show spanning-tree" in cmd_lc:
                ridx = low.index("role")  if "role"  in low else None
                sidx = low.index("state") if "state" in low else None
                root = blocking = forwarding = 0
                for r in rows:
                    role = str(r[ridx]).strip().lower()  if (ridx is not None and ridx < len(r) and r[ridx]) else ""
                    st   = str(r[sidx]).strip().lower()   if (sidx is not None and sidx < len(r) and r[sidx]) else ""
                    if role == "root":        root += 1
                    if st == "blocking":      blocking += 1
                    if st == "forwarding":    forwarding += 1
                write_resumo_block(ws, "Resumo – Spanning-Tree",
                                   [("ROOT ports", root), ("FORWARDING", forwarding), ("BLOCKING", blocking)],
                                   start_row=end_row + 2)

        else:
            # Sem grid → despejo do texto bruto (com formatação de monospace)
            ws.cell(row=start_row - 1, column=1, value="Output (texto bruto):").font = Font(bold=True)
            lines = (raw.splitlines() if isinstance(raw, str) else [])
            for i, line in enumerate(lines, start=start_row):
                c = ws.cell(row=i, column=1, value=line)
                c.alignment = wrap
                c.font = mono
            ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 100)

    # Dashboard sempre no fim (todas as folhas já existem)
    build_dashboard(wb)
    wb.save(xlsx_path)

